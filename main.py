from openpyxl import load_workbook

wb = load_workbook("textspel.xlsx")
ws = wb.active
room_number = 2
room_number_change_n = -1
room_number_change_s = 1
room_position_change = 1
#Den visar hur långt spelaren rör sig
position = {1:'A', 2:'B', 3:'C', 4:"D"}
room_position = 3
current_room = 2
start = True
wsRooms = wb["Rooms"]
wsMap = wb["Map"]
save_file = wb['Save']


#denna sparar min position i excel dokumentet.
wb.save("textspel.xlsx")
room_number = (int(str(save_file["A2"].value)))




def look():
    print(ws[F"{position[room_position]}{room_number}"].value)
    #Printar ut value på vilken cell och vilker rad du är på


def north():
    global room_number
    if room_number >= 1:
        room_number += room_number_change_n
        print(ws[F"{position[room_position]}{room_number}"].value)
        # Den kommer att ta room_number och adera plus -1 för att röra den ett steg uppåt i dokumentet

def south():
    global room_number
    if room_number >= 1:
        room_number += room_number_change_s
        print(ws[F"{position[room_position]}{room_number}"].value)
        #Den kommer att ta room_number och adera plus 1 för att röra den ett steg neråt i dokumentet


def west():
    global room_position
    if room_position >= 1:
        room_position -= room_position_change
        print(ws[F"{position[room_position]}{room_number}"].value)
        #Gör exakt samma sak som East def fast den subtraherar istället för addera

def east():
    global room_position
    if room_position >= 1:
        room_position += room_position_change
        print(ws[F"{position[room_position]}{room_number}"].value)
        #Jag har gjort bokstäverna till siffror.
        #Det gör att jag kan addera eller subtrahera ett tal för att förflytta spelaren.



def help():
    print ("look,north,south,west,stop,active room")
    #Den printar alla våra commands




while start:
    #Det är en input som gör det möjligt att få våra outputs
    #Om en inout finns så kommer den kalla på respektive define
    reading = input("")
    if reading == "look":
        look()
    if reading == 'north':
        north()
    if reading == "south":
        south()
    if reading == "west":
        west()
    if reading == "east":
        east()
    elif reading == "stop":
        start = False
    if reading == 'quit':
        start = False
    if reading == "help":
        help()
    if reading == "save":
        save_file["A2"] = room_number
        wb.save("textspel.xlsx")

